import numpy as np
import random
import time
import os
import glob
import sys
import cv2
import imageio
from scipy import misc, ndimage, io, linalg, stats
from PIL import Image, ImageOps
from time import gmtime
import tensorflow as tf
from tensorflow.keras.callbacks import ModelCheckpoint
from tensorflow.keras.models import Model, load_model
from tensorflow.keras.layers import Input, Conv2D, MaxPooling2D, UpSampling2D, concatenate, BatchNormalization
from tensorflow.keras.optimizers import Adam
from tensorflow.keras import losses
import h5py
import shutil
from scipy.ndimage.filters import gaussian_filter
import itertools
import queue
import multiprocessing as mp
import matplotlib.pyplot as plt
from itertools import cycle
import copy
import math
import skimage.measure
import seaborn as sns
import sklearn
import subprocess
import datetime
import gc
sns.set_style('ticks')
# the path should point to the FIt-SNE directory
import sys; sys.path.append('../../FIt-SNE')
import openpyxl
from constants import *
import warnings

warnings.simplefilter("ignore") # Change the filter in this process
os.environ["PYTHONWARNINGS"] = "ignore" # Also affect subprocesses


Image.MAX_IMAGE_PIXELS = 933120000

# creates as many fold folders as we need
# (each fold will reserve 1 image for validation out of the set and use the rest for training)
# same for 'images' folder
# folds folder will store the preprocessed data, trained model files, and results
def make_fold_folders(num_folds):
    num_folds += 1
    count_folds = 1
    folds_folder = os.path.join(os.getcwd(), 'folds')
    while count_folds <= num_folds:
        curr_fold = os.path.join(folds_folder, 'f' + str(count_folds))
        os.mkdir(curr_fold)
        count_exp = 1

        log = os.path.join(curr_fold, 'log')
        os.mkdir(log)
        model = os.path.join(curr_fold, 'model')
        os.mkdir(model)
        data = os.path.join(curr_fold, 'preprocessed_data')
        os.mkdir(data)
        results = os.path.join(curr_fold, 'results')
        os.mkdir(results)
        while count_exp <= num_folds:
            os.mkdir(os.path.join(log, 'exp' + str(count_exp)))
            os.mkdir(os.path.join(model, 'exp' + str(count_exp)))
            os.mkdir(os.path.join(data, 'exp' + str(count_exp)))
            os.mkdir(os.path.join(data, 'exp' + str(count_exp), 'training_data'))
            os.mkdir(os.path.join(data, 'exp' + str(count_exp), 'validation_data'))
            os.mkdir(os.path.join(results, 'exp' + str(count_exp)))
            count_exp += 1
        count_folds += 1
    os.rename(os.path.join(folds_folder, 'f' + str(num_folds)), os.path.join(folds_folder, 'f_tsne'))


# computes the dice validation score between a prediction and the actual
def dice(bm_pred, bm_val):
    intersection = np.multiply(bm_pred, bm_val)
    seg_cropount = np.sum(bm_pred)
    ref_count = np.sum(bm_val)
    int_count = np.sum(intersection)
    dice_similarity = 2 * int_count / (ref_count + seg_cropount)
    return dice_similarity

# creates 256x256 patch data for each image
def prep_image(fold, exp, img_name, numerator, patch_size, patches_per_image, cpu):

    image_folder = os.path.join(os.getcwd(), 'images', fold, 'train_and_val')
    #multiplying 256 by sqrt(2) to allow for crop after rotation
    prerotated_patch_size = int(patch_size * 1.4142)
    trn_count = patches_per_image - numerator
    # number of passes
    k = 0

    #to seed random differently for each process
    random_state = np.random.RandomState()

    # each batch is an array of patches from a single image
    img_trn_patches = np.ndarray((trn_count, patch_size, patch_size, 1), dtype='float32')
    seg_trn_patches = np.ndarray((trn_count, patch_size, patch_size, 1), dtype='float32')
    img_val_patches = np.ndarray((numerator, patch_size, patch_size, 1), dtype='float32')
    seg_val_patches = np.ndarray((numerator, patch_size, patch_size, 1), dtype='float32')

    # region of interest
    roi_file = img_name + '_roi.png'
    roi = Image.open(os.path.join(image_folder, img_name, roi_file)).convert("L")
    roi = np.array(roi)
    # original image
    img_file = img_name + '_img.png'
    #extracting green channel
    img = Image.open(os.path.join(image_folder, img_name, img_file)).convert("RGB")
    red, img, blue = img.split()
    img = np.array(img)
    # segmentation
    seg_file = img_name + '_seg.png'
    seg = Image.open(os.path.join(image_folder, img_name, seg_file)).convert("L")
    seg = np.array(seg)
    # because the red segmentation color is below 127.5 so set to 0 otherwise
    seg[np.nonzero(seg)] = 255

    # converting to binary
    roi[roi < 0] = 0
    roi[(roi > 0) & (roi <= 127.5)] = 0
    roi[(roi > 127.5)] = 1
    seg[seg < 0] = 0
    seg[(seg > 0) & (seg <= 127.5)] = 0
    seg[(seg > 127.5)] = 1
    seg[roi == 0] = 0

    # approximating length of region to square
    A = len(np.nonzero(roi)[0])
    D = np.sqrt(A)
    Ct = ndimage.measurements.center_of_mass(roi)
    # finding bounding points
    x1 = int(Ct[0] - D)
    x2 = int(Ct[0] + D)
    y1 = int(Ct[1] - D)
    y2 = int(Ct[1] + D)
    # ensuring bounding point is not off image
    if x1 < 0:
        x1 = 0
    if x2 > roi.shape[0]:
        x2 = roi.shape[0]
    if y1 < 0:
        y1 = 0
    if y2 > roi.shape[1]:
        y2 = roi.shape[1]
    # cutting out square region
    roi_c = roi[x1:x2, y1:y2]
    roi_t = roi_c

    # constructing augmented patches
    # meshgrid of possible patch locations
    x = np.arange(0.75 * prerotated_patch_size, roi_t.shape[0] - 0.75 * prerotated_patch_size, 20)
    y = np.arange(0.75 * prerotated_patch_size, roi_t.shape[1] - 0.75 * prerotated_patch_size, 20)

    xx, yy = np.meshgrid(x, y, sparse=False)
    xx = np.resize(xx, (np.product(xx.shape)))
    yy = np.resize(yy, (np.product(yy.shape)))


    kk = list(range(np.product(yy.shape)))

    # constructing the patches now
    while k < patches_per_image:
        # pick a random patch location
        kc = random_state.choice(kk)
        xc = xx[kc]
        yc = yy[kc]
        # dimension of the bounding box
        window_size = prerotated_patch_size
        x_min = int(xc - window_size / 2)
        x_max = int(xc + window_size / 2)
        y_min = int(yc - window_size / 2)
        y_max = int(yc + window_size / 2)
        # cutting patch out of roi
        roi_p = roi_t[x_min:x_max, y_min:y_max]
        roi_ratio = np.count_nonzero(roi_p) / (roi_p.shape[0] * roi_p.shape[1])

        # if entire patch is yellow
        if (random_state.rand(1) > 0.9 and roi_ratio >= ROI_RATIO) or (random_state.rand(1) <= 0.9 and np.count_nonzero(roi_p) / (roi_p.shape[0] * roi_p.shape[1]) == 1):
            # section of image within roi bounding box
            img_crop = img[x1:x2, y1:y2]
            # section of seg within roi bounding box
            seg_crop = seg[x1:x2, y1:y2]

            #rotating larger patch
            # rotates image by same factor that roi was rotated by
            rotated_img = Image.fromarray(img_crop[x_min:x_max, y_min:y_max])
            ang_rotate = random_state.randint(0, 360)
            rotated_img_arr = np.array(rotated_img.rotate(ang_rotate))
            # rotates seg by same factor
            rotated_seg = Image.fromarray(seg_crop[x_min:x_max, y_min:y_max])
            rotated_seg_arr = np.array(rotated_seg.rotate(ang_rotate))

            # elastic deformation?
            #np.resize(rotated_img_arr, (prerotated_patch_size, prerotated_patch_size))
            #np.resize(rotated_seg_arr, (prerotated_patch_size, prerotated_patch_size))

            #cropping larger, now rotated patch to correct size
            patch_center = prerotated_patch_size // 2
            img_patch = rotated_img_arr[patch_center - patch_size // 2:patch_center + patch_size // 2, patch_center - patch_size // 2:patch_center + patch_size // 2]
            seg_patch = rotated_seg_arr[patch_center - patch_size // 2:patch_center + patch_size // 2, patch_center - patch_size // 2:patch_center + patch_size // 2]
            #Image.fromarray(seg_patch * 255).convert("L").show()

            #shadows
            draw = random_state.uniform(0, 1)
            if draw > SHADOW:
                copy_patch = img_patch.copy()
                filler = random_state.randint(1, 10)
                mode = stats.mode(copy_patch, axis=None)[0]
                if draw > 0.96:
                    copy_patch = np.hstack((mode * np.ones((patch_size, filler)), copy_patch))
                    copy_patch = random_state.uniform(0.125, 0.375) * copy_patch[:, :patch_size]
                    img_patch = copy_patch + img_patch
                else:
                    copy_patch = np.hstack((copy_patch, mode * np.ones((patch_size, filler))))
                    copy_patch = random_state.uniform(0.125, 0.375) * copy_patch[:, filler:]
                    img_patch = copy_patch + img_patch


            ##adding bubbles
            if random_state.rand(1) > BUBBLE:
                for i in range(random_state.randint(0, (1 * patch_size // 256)**2 + 1)):
                    cv2.circle(img_patch, (random_state.randint(20, patch_size - 20), random_state.randint(20, patch_size - 20)), random.randint(0, 10), (0, 0, 0), thickness=np.random.randint(2, 4), lineType=cv2.LINE_AA, shift=0)

            #multivariate normal
            img_patch = img_patch.astype('int16')

            #gaussian noise
            if random_state.rand(1) > G_NOISE:

                multivar = stats.multivariate_normal(mean=[random_state.randint(5, patch_size - 5), random_state.randint(5, patch_size - 5)],
                           cov=[[np.random.randint(3000, 4500), 0], [0, np.random.randint(3000, 4500)]], seed = random_state)
                x, y = np.mgrid[0:patch_size, 0:patch_size]
                pos = np.dstack((x, y))
                multivar_pix = multivar.pdf(pos)
                multivar_max = np.amax(multivar_pix)
                multivar  = np.array(multivar_pix)

                multivar = multivar / multivar_max * random_state.randint(50, 70)
                if random_state.rand(1) > 0.5:
                    img_patch = img_patch + multivar * 1
                else:
                    img_patch = img_patch - multivar * 1
                #case where wraparound above 255 occurs
                img_patch[img_patch > 255] = 255
                #wraparound below 255
                img_patch[img_patch < 0] = 0

            #adjusting brightness
            img_patch = np.add(img_patch, random_state.randint(-BRIGHT_MAG, BRIGHT_MAG))
            #case where wraparound above 255 occurs
            img_patch[img_patch > 255] = 255
            #wraparound below 255
            img_patch[img_patch < 0] = 0

            img_patch = img_patch.astype('uint8')

            #flipping
            if random_state.rand(1) > FLIP_UD:
                np.flipud(img_patch)
                np.flipud(seg_patch)

            if random_state.rand(1) > FLIP_LR:
                np.fliplr(img_patch)
                np.fliplr(seg_patch)

            #gaussian blur
            img_patch = gaussian_filter(img_patch, sigma=random_state.randint(0, 3))

            #fourier transform taken from http://www.imagemagick.org/Usage/fourier/#contrast
            if random_state.rand(1) > FT_PROB:
                transformed_patch = np.fft.fft2(img_patch)
                contrasted_patch = linalg.fractional_matrix_power(transformed_patch, random_state.uniform(0.95, 1.02))
                img_patch = np.abs(np.fft.ifft2(contrasted_patch))

            img_patch = Image.fromarray(img_patch).convert('L')

            #CLAHE if necessary
            if "CLAHE" not in img_name:
                x = random_state.uniform(0, 0.02)
                img_patch = ImageOps.autocontrast(img_patch, cutoff = x)

            img_patch = np.array(img_patch)



            # converting to binary
            seg_patch[seg_patch < 0] = 0
            seg_patch[(seg_patch > 0) & (seg_patch <= 0.5)] = 0
            seg_patch[(seg_patch > 0.5)] = 1

            # patch is added to array
            if k < trn_count:
                img_trn_patches[k, ...] = img_patch[np.newaxis, ..., np.newaxis]
                seg_trn_patches[k, ...] = seg_patch[np.newaxis, ..., np.newaxis]

            else:
                img_val_patches[k - trn_count, ...] = img_patch[np.newaxis, ..., np.newaxis]
                seg_val_patches[k - trn_count, ...] = seg_patch[np.newaxis, ..., np.newaxis]
            k += 1
    save_dir = os.path.join(os.getcwd(), 'folds', fold, 'preprocessed_data', exp, img_name)
    if not os.path.isdir(save_dir):
        try:
            os.mkdir(save_dir)
        except FileExistsError:
            pass
    np.save(os.path.join(save_dir, img_name + 'img_trn_patches' + str(cpu) + '.npy'), img_trn_patches)
    np.save(os.path.join(save_dir, img_name + 'img_val_patches' + str(cpu) + '.npy'), img_val_patches)
    np.save(os.path.join(save_dir, img_name + 'seg_trn_patches' + str(cpu) + '.npy'), seg_trn_patches)
    np.save(os.path.join(save_dir, img_name + 'seg_val_patches' + str(cpu) + '.npy'), seg_val_patches)


# we are preprocessing for both the training and validation datasets at the same time
def prep(fold, exp, epoch, trn_imgs, numerator, patch_size, patches_per_image):  # takes in fold, exp, cwd
    # assuming this script is run from the main folder


    cwd = os.getcwd()

    data_path = os.path.join(cwd, 'data' + str(patch_size))
    if not os.path.exists(data_path):
        os.mkdir(data_path)

    path = os.path.join(cwd, 'data' + str(patch_size), 'epoch' + str(epoch))
    if os.path.exists(path):
        return

    #numerator / patches_per_image will yield the percent of validation split
    trn_count = patches_per_image - numerator
    img_trn_data = np.ndarray(((trn_count) * trn_imgs.__len__(), patch_size, patch_size, 1), dtype='float32')
    seg_trn_data = np.ndarray(((trn_count) * trn_imgs.__len__(), patch_size, patch_size, 1), dtype='float32')
    img_val_data = np.ndarray((numerator * trn_imgs.__len__(), patch_size, patch_size, 1), dtype='float32')
    seg_val_data = np.ndarray((numerator * trn_imgs.__len__(), patch_size, patch_size, 1), dtype='float32')
    print(img_trn_data.shape)

    image_folder = os.path.join(cwd, 'images', fold, 'train_and_val')
    # Clahe'd stuff
    sequence_name = str(1)
    time = gmtime()
    print('No. ', sequence_name, ' starting time: ', str(time.tm_hour), ':', str(time.tm_min), ':', str(time.tm_sec))
    # number of samples processed
    ks = 0
    trn_imgs = [x for x in trn_imgs if x[0] != '.']
    #use multiprocessing on each image

    #process for each image
    img_files = os.listdir(image_folder)
    num_images = len(img_files)
    p = {}

    #creates a new process for each image and assigns it to number in dictionary
    for img_num in range(num_images):
        img_name = img_files[img_num]
        for cpu in range(available_cpus):
            p[cpu] = mp.Process(target=prep_image, args=(fold, exp, img_name, numerator // available_cpus, patch_size, patches_per_image // available_cpus, cpu))
            p[cpu].start()
            print("starting process")
            #tells python to wait until each process is finished before continuing
        for cpu in range(available_cpus):
            p[cpu].join()
        data_folder = os.path.join(os.getcwd(), 'folds', fold, 'preprocessed_data', exp, img_name)
        all_patches = os.listdir(data_folder)
        img_trn = [i for i in all_patches if 'img_trn' in i]
        img_val = [i for i in all_patches if 'img_val' in i]
        seg_trn = [i for i in all_patches if 'seg_trn' in i]
        seg_val = [i for i in all_patches if 'seg_val' in i]
        #so that corresponding img and seg files match up
        img_trn.sort()
        img_val.sort()
        seg_trn.sort()
        seg_val.sort()
        img_trn = [np.load(os.path.join(data_folder, i)) for i in img_trn]
        img_val = [np.load(os.path.join(data_folder, i)) for i in img_val]
        seg_trn = [np.load(os.path.join(data_folder, i)) for i in seg_trn]
        seg_val = [np.load(os.path.join(data_folder, i)) for i in seg_val]
        img_trn_patches = np.concatenate(img_trn, axis = 0)
        img_val_patches = np.concatenate(img_val, axis = 0)
        seg_trn_patches = np.concatenate(seg_trn, axis = 0)
        seg_val_patches = np.concatenate(seg_val, axis = 0)
        np.save(os.path.join(data_folder, img_name + 'img_trn_patches.npy'), img_trn_patches)
        np.save(os.path.join(data_folder, img_name + 'img_val_patches.npy'), img_val_patches)
        np.save(os.path.join(data_folder, img_name + 'seg_trn_patches.npy'), seg_trn_patches)
        np.save(os.path.join(data_folder, img_name + 'seg_val_patches.npy'), seg_val_patches)


    print("processing done")
    for ks in range(1, num_images + 1):
        img_name = img_files[ks - 1]
        data_folder = os.path.join(os.getcwd(), 'folds', fold, 'preprocessed_data', exp, img_name)

        for filename in os.listdir(data_folder):
            patches = np.load(os.path.join(data_folder, filename))

            if 'img_trn_patches.npy' in filename:
                img_trn_data[(ks - 1) * trn_count:ks * trn_count] = patches
            if 'seg_trn_patches.npy' in filename:
                seg_trn_data[(ks - 1) * trn_count:ks * trn_count] = patches
            if 'img_val_patches.npy' in filename:
                img_val_data[(ks - 1) * numerator:ks * numerator] = patches
            if 'seg_val_patches.npy' in filename:
                seg_val_data[(ks - 1) * numerator:ks * numerator] = patches
        shutil.rmtree(data_folder)
    print("concatenation done")
    time = gmtime()
    print('No. ', sequence_name, ' ending time: ', str(time.tm_hour), ':', str(time.tm_min), ':', str(time.tm_sec))

    save_dir = os.path.join(cwd, 'folds', fold, 'preprocessed_data', exp)
    trn_save_dir = os.path.join(save_dir, 'training_data')
    val_save_dir = os.path.join(save_dir, 'validation_data')


    os.mkdir(path)
    np.save(os.path.join(path, 'img_trn_data.npy'), img_trn_data)
    np.save(os.path.join(path, 'seg_trn_data.npy'), seg_trn_data)
    np.save(os.path.join(path, 'img_val_data.npy'), img_val_data)
    np.save(os.path.join(path, 'seg_val_data.npy'), seg_val_data)

    return

#separates images into folds
def separate_test(num_folds, num_test_images):

    cwd = os.getcwd()
    all_images_folder = os.path.join(cwd, 'images', 'all_images')
    all_image_names = os.listdir(all_images_folder)
    all_image_names = [x for x in all_image_names if x[0] != '.']
    images_seen = []
    comb = list(itertools.combinations(all_image_names, num_test_images))



    for fold in range(num_folds):
        os.mkdir(os.path.join(cwd, 'images', 'f' + str(fold + 1)))
        os.mkdir(os.path.join(cwd, 'images', 'f' + str(fold + 1), 'test'))
        os.mkdir(os.path.join(cwd, 'images', 'f' + str(fold + 1), 'train_and_val'))

        while len(images_seen) == fold:
            image_combination = comb[random.randint(0, len(comb) - 1)] #array of strings
            unique = True
            for fold_comb_index in range(len(images_seen)):
                if image_combination == images_seen[fold_comb_index]:
                    unique = False
            if unique:
                images_seen.append(image_combination)
        for image_name in all_image_names:
            if image_name in images_seen[-1]:
                shutil.copytree(os.path.join(all_images_folder, image_name), os.path.join(cwd, 'images', 'f' + str(fold + 1), 'test', image_name))
            else:
                shutil.copytree(os.path.join(all_images_folder, image_name), os.path.join(cwd, 'images', 'f' + str(fold + 1), 'train_and_val', image_name))

    os.mkdir(os.path.join(cwd, 'images', 'f_tsne'))
    os.mkdir(os.path.join(cwd, 'images', 'f_tsne', 'test'))
    os.mkdir(os.path.join(cwd, 'images', 'f_tsne', 'train_and_val'))
    for i in os.listdir(all_images_folder):
        shutil.copytree(os.path.join(all_images_folder, i), os.path.join(cwd, 'images', 'f_tsne', 'train_and_val', i))



#Main Train function
def train(fold, exp, init_val, end_val, numerator, patch_size, patches_per_image, gpu=0):

    os.environ['CUDA_VISIBLE_DEVICES'] = str(gpu)
    trn_count = patches_per_image - numerator
    cwd = os.getcwd()
    test_folder = os.path.join(cwd, 'images', fold, 'test')
    test_case = os.listdir(test_folder)[0]
    test_folder = os.path.join(test_folder, test_case)
    #compute on seg
    for k_init in range(init_val, end_val, 1):
        trn_imgs = os.listdir(os.path.join(cwd, 'images', fold, 'train_and_val'))

        prep(fold, exp, k_init + 1, trn_imgs, numerator, patch_size, patches_per_image)

        # training data
        trn_imgs = os.listdir(os.path.join(cwd, 'images', fold, 'train_and_val'))

        print('Epoch:', str(k_init + 1))
        # number of epochs
        k_init += 1
        trn_data = np.ndarray((trn_imgs.__len__() * trn_count, patch_size, patch_size, 1), dtype='float32')
        seg_trn_data = np.ndarray((trn_imgs.__len__() * trn_count, patch_size, patch_size, 1), dtype='float32')
        # going through each original image
        #FIXME This outer for loop is unnecessary
        #FIXME Also we only need to load in the data directly now, like how we did for img_val after for loop
        #for test_case in trn_imgs:
            #sample_folder = os.path.join(data_folder, sample_train)
            #folder_content = os.listdir(sample_folder)
            #img_list = [x for x in folder_content if 'img' in x]
            #img_sample = img_list[k_init]
            #seg_sample = 'seg' + img_sample[3:]
            #img_trn_data = np.load(os.path.join(sample_folder, img_sample))
            #seg_trn_data = np.load(os.path.join(sample_folder, seg_sample))
            #img_trn[(ks - 1) * patches_per_image:ks * patches_per_image, ...] = img_trn_data
            #seg_trn[(ks - 1) * patches_per_image:ks * patches_per_image, ...] = seg_trn_data
            #img_val = np.load(os.path.join(val_dir, exp, 'img_val_1.npy'))
            #seg_val = np.load(os.path.join(val_dir, exp, 'seg_val_1.npy'))
            #file_list = os.listdir(data_folder)
            #for file in file_list:
            #    data = np.load(os.path.join(data_folder, file))
            #    if 'img' in file:
            #        img_count += 1
            #        trn_data[(img_count - 1) * trn_count:img_count * trn_count, ...] = data
            #    else:
            #        seg_count += 1
            #        seg_trn_data[(seg_count - 1) * trn_count:seg_count * trn_count, ...] = data
        trn_data = np.load(os.path.join(cwd, 'data' + str(patch_size), 'epoch' + str(k_init), 'img_trn_data.npy'))
        seg_trn_data = np.load(os.path.join(cwd, 'data' + str(patch_size), 'epoch' + str(k_init), 'seg_trn_data.npy'))
        img_val = np.load(os.path.join(cwd, 'data' + str(patch_size), 'epoch' + str(k_init), 'img_val_data.npy'))
        seg_val = np.load(os.path.join(cwd, 'data' + str(patch_size), 'epoch' + str(k_init), 'seg_val_data.npy'))

        # shuffling training data and seg data randomly, but so that they are still in relative order
        c = list(zip(trn_data, seg_trn_data))
        np.random.shuffle(c)
        trn_data, seg_trn_data = zip(*c)

        img_train = np.concatenate((trn_data, img_val), axis=0)
        seg_train = np.concatenate((seg_trn_data, seg_val), axis=0)
        val_split = img_val.__len__() / img_train.__len__()

        model_folder = os.path.join(cwd, 'folds', fold, 'model', exp)
        model_list = os.listdir(model_folder)

        logdir = "logs/fit/" + datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
        tensorboard_callback = tf.keras.callbacks.TensorBoard(log_dir=logdir)

        #mirrored_strategy = tf.distribute.MirroredStrategy()
        # only for first model
        if model_list.__len__() == 0:
            patch_rows = patch_size
            patch_cols = patch_size
            #with mirrored_strategy.scope():
            inputs = Input((patch_rows, patch_cols, 1))
            conv1 = Conv2D(32, (3, 3), activation='relu', padding='same')(inputs)
            conv1 = BatchNormalization()(conv1)
            conv1 = Conv2D(32, (3, 3), activation='relu', padding='same')(conv1)
            conv1 = BatchNormalization()(conv1)
            pool1 = MaxPooling2D(pool_size=(2, 2))(conv1)

            conv2 = Conv2D(64, (3, 3), activation='relu', padding='same')(pool1)
            conv2 = BatchNormalization()(conv2)
            conv2 = Conv2D(64, (3, 3), activation='relu', padding='same')(conv2)
            conv2 = BatchNormalization()(conv2)
            pool2 = MaxPooling2D(pool_size=(2, 2))(conv2)

            conv3 = Conv2D(128, (3, 3), activation='relu', padding='same')(pool2)
            conv3 = BatchNormalization()(conv3)
            conv3 = Conv2D(128, (3, 3), activation='relu', padding='same')(conv3)
            conv3 = BatchNormalization()(conv3)
            pool3 = MaxPooling2D(pool_size=(2, 2))(conv3)

            conv4 = Conv2D(256, (3, 3), activation='relu', padding='same')(pool3)
            conv4 = BatchNormalization()(conv4)
            conv4 = Conv2D(256, (3, 3), activation='relu', padding='same')(conv4)
            conv4 = BatchNormalization()(conv4)
            pool4 = MaxPooling2D(pool_size=(2, 2))(conv4)

            conv5 = Conv2D(512, (3, 3), activation='relu', padding='same')(pool4)
            conv5 = BatchNormalization()(conv5)
            conv5 = Conv2D(512, (3, 3), activation='relu', padding='same')(conv5)
            conv5 = BatchNormalization()(conv5)

            up6 = concatenate([UpSampling2D(size=(2, 2))(conv5), conv4], axis=3)
            conv6 = Conv2D(256, (3, 3), activation='relu', padding='same')(up6)
            conv6 = BatchNormalization()(conv6)
            conv6 = Conv2D(256, (3, 3), activation='relu', padding='same')(conv6)
            conv6 = BatchNormalization()(conv6)

            up7 = concatenate([UpSampling2D(size=(2, 2))(conv6), conv3], axis=3)
            conv7 = Conv2D(128, (3, 3), activation='relu', padding='same')(up7)
            conv7 = BatchNormalization()(conv7)
            conv7 = Conv2D(128, (3, 3), activation='relu', padding='same')(conv7)
            conv7 = BatchNormalization()(conv7)

            up8 = concatenate([UpSampling2D(size=(2, 2))(conv7), conv2], axis=3)
            conv8 = Conv2D(64, (3, 3), activation='relu', padding='same')(up8)
            conv8 = BatchNormalization()(conv8)
            conv8 = Conv2D(64, (3, 3), activation='relu', padding='same')(conv8)
            conv8 = BatchNormalization()(conv8)

            up9 = concatenate([UpSampling2D(size=(2, 2))(conv8), conv1], axis=3)
            conv9 = Conv2D(32, (3, 3), activation='relu', padding='same')(up9)
            conv9 = BatchNormalization()(conv9)
            conv9 = Conv2D(32, (3, 3), activation='relu', padding='same')(conv9)
            conv9 = BatchNormalization()(conv9)

            conv10 = Conv2D(1, (1, 1), activation='sigmoid')(conv9)

            model = Model(inputs=[inputs], outputs=[conv10])

            epoch_number = 0

        # load the last saved model
        else:
            # last save
            # model_list contains the names of the model after each epoch
            # ^ is a listdir of model_folder
            # below could be commented out
            model_file = model_list[-1]
            model_idx_list = []
            # looping through each model name
            for model_sample in model_list:
                # getting the digits of the model name
                model_idx = int(model_sample[8:10])
                model_idx_list.append(model_idx)
            epoch_number = int(np.max(model_idx_list))
            if epoch_number < 10:
                # make two-digit
                model_file_ref = model_file[0:8] + '0' + str(epoch_number) + model_file[10:11]
            else:
                model_file_ref = model_file[0:8] + str(epoch_number) + model_file[10:11]
            for model_sample in model_list:
                if model_file_ref in model_sample:
                    model_file = model_sample
            # reading model file
            model = load_model(os.path.join(model_folder, model_file))
        #with mirrored_strategy.scope():
        model.compile(optimizer=Adam(lr=learning_rate), loss=losses.binary_crossentropy, metrics=['accuracy'])

        if k_init < 10:
            file_name = 'weights.' + '0' + str(k_init) + '-{val_loss:.3f}.h5'
        else:
            file_name = 'weights.' + str(k_init) + '-{val_loss:.3f}.h5'
        model_check_file = os.path.join(model_folder, file_name)

        # specifies location where the new model is to be saved
        # may be a newer approach
        model_checkpoint = ModelCheckpoint(model_check_file, monitor='val_loss', save_best_only=False)

        model.fit(img_train, seg_train, batch_size=BATCH_SIZE, epochs=1, verbose=1, shuffle=True, validation_split=val_split,
                  callbacks=[model_checkpoint, tensorboard_callback])
        print(fold, test_folder, test_case, model_check_file, epoch_number)
        #list_of_files = glob.glob(model_folder + '/*') # * means all if need specific format then *.csv
        #latest_model_file = max(list_of_files, key=os.path.getctime)
        #print(latest_model_file)
        #subprocess.call(['sh', './calc_leaf_metrics_for_epoch.sh', fold, test_case, test_case, "False", str(k_init)])
        print("saved")
        #cuda.select_device(gpu)
        #cuda.close()

    #dice_validation(fold, exp)
    #plot_vein_stats(fold, test_case, k_init)
    return

#Plotting vein density stats from excel file
def plot_vein_stats(fold, test_case, epochs):
    vein_density = []
    vein_loopiness = []
    seg_density = 0
    seg_loopiness = 0
    for epoch in range(0, epochs + 1):
        path = os.path.join(os.getcwd(), 'images', fold, 'test', test_case)
        ps = openpyxl.load_workbook(os.path.join(path, 'results_table_epoch_' + str(epoch) + '.xlsx'))
        sheet = ps['Sheet1']
        density_val = sheet['H2'].value
        loopiness_val = sheet['AH2'].value
        if epoch == 0:
            seg_density = density_val
            seg_loopiness = loopiness_val
        else:
            vein_density.append(density_val)
            vein_loopiness.append(loopiness_val)

    plt.figure()
    plt.title('Vein Density over Epochs')
    plt.plot(range(1, epochs + 1), vein_density, label='Model density')
    plt.plot(range(1, epochs + 1), [seg_density] * len(range(1, epochs + 1)), label='Real density')
    plt.legend()
    plt.savefig(path + '/vein_density.png')

    plt.figure()
    plt.title('Vein Loopiness over Epochs')
    plt.plot(range(1, epochs + 1), vein_loopiness, label='Model loopiness')
    plt.plot(range(1, epochs + 1), [seg_loopiness] * len(range(1, epochs + 1)), label='Real loopiness')
    plt.legend()
    plt.savefig(path + '/vein_loopiness.png')

    return

# similarity between the output seg image and hand-traced seg image
def dice_validation(fold, exp):
    # all patches
    cwd = os.getcwd()
    val_dir = os.path.join(cwd, 'folds', fold, 'preprocessed_data', exp, 'validation_data')
    img_val_set = np.load(os.path.join(val_dir, 'img_val_data.npy'))
    seg_val_set = np.load(os.path.join(val_dir, 'seg_val_data.npy'))

    model_folder = os.path.join(cwd, 'folds', fold, 'model', exp)
    model_list_init = os.listdir(model_folder)
    epoch_number = 0
    model_file_init = model_list_init[0]
    model_list = []
    # hopefully clarified by 291-311
    for model_num in range(len(model_list_init)):
        epoch_number += 1
        if epoch_number < 10:
            str_cmp = str(model_file_init[0:8] + '0' + str(epoch_number))
            model_file = [s for s in model_list_init if str_cmp in s]
        else:
            str_cmp = str(model_file_init[0:8] + str(epoch_number))
            model_file = [s for s in model_list_init if str_cmp in s]
        print(model_file)
        model_list.append(str(model_file[0]))

    D = []
    km = 0
    # for each model, predicts the seg output vs hand-traced output and appends dice score to array D
    for model_file in model_list:
        km += 1
        print('Progress: ', km, '/', len(model_list))
        model = load_model(os.path.join(model_folder, model_file))

        prd_val_set = model.predict(img_val_set, batch_size=16)
        prd_val_set[prd_val_set >= 0.5] = 1
        prd_val_set[prd_val_set < 0.5] = 0
        d = dice(prd_val_set, seg_val_set)
        D.append(d)
    print(D)
    save_dir = os.path.join(cwd, 'folds', fold, 'log', exp)
    np.save(os.path.join(save_dir, 'dice_validation.npy'), D)
    return


#test some fold
def f_test(fold, exp):
    cwd = os.getcwd()
    #grab model
    #divide test image into patches
    #predict on test image
    #show output
    test_folder = os.path.join(os.path.join(cwd, 'images', fold, 'test'))
    test_sample = os.listdir(test_folder)
    #we probably have to make an array of these

    model_folder = os.path.join(cwd, 'folds', fold, 'model', exp)
    D_list = np.load(os.path.join(os.path.join(cwd, 'folds', fold, 'log', exp, 'dice_validation.npy')))
    D_list_copy = D_list.copy()
    epsilon = 0.000001
    for i in range(len(D_list)):
        D_list += epsilon * i
    max_dice_index = np.argmax(D_list)

    model_list_init = os.listdir(model_folder)
    epoch_number = 0
    model_file_init = model_list_init[0]
    model_list = []
    # hopefully clarified by 291-311
    for model_num in range(len(model_list_init)):
        epoch_number += 1
        if epoch_number < 10:
            str_cmp = str(model_file_init[0:8] + '0' + str(epoch_number))
            model_file = [s for s in model_list_init if str_cmp in s]
        else:
            str_cmp = str(model_file_init[0:8] + str(epoch_number))
            model_file = [s for s in model_list_init if str_cmp in s]
        model_list.append(str(model_file[0]))

    model_file = model_list[max_dice_index]
    max_dice = []
    max_dice.append(model_file)
    print(model_file)
    model = load_model(os.path.join(model_folder, model_file))

    result_save_folder = os.path.join(cwd, 'folds', fold, 'results', exp)
    np.save(os.path.join(result_save_folder, 'max_dice.npy'), np.array(max_dice))

    count = 0
    for test_case in test_sample:
        count += 1
        print('Test case: ' + test_case + 'Progress: ' + str(count) + ' / ' + str(len(test_sample)))
        if not os.path.exists(os.path.join(result_save_folder, test_case)):
            os.mkdir(os.path.join(result_save_folder, test_case))
        img_test_sample = test_case + '_img.png'
        roi_test_sample = test_case + '_roi.png'
        seg_test_sample = test_case + '_seg.png'
        # region of interest
        roi = Image.open(os.path.join(cwd, test_folder, test_case, roi_test_sample)).convert("L")
        roi = np.array(roi)
        # original image
        img = Image.open(os.path.join(cwd, test_folder, test_case, img_test_sample)).convert("L")
        img = ImageOps.autocontrast(img, cutoff = 0.01)
        img = np.array(img)
        # segmentation
        seg = Image.open(os.path.join(cwd, test_folder, test_case, seg_test_sample)).convert("L")
        seg = np.array(seg)
        # because the red segmentation color is below 127.5 so set to 0 otherwise
        seg[np.nonzero(seg)] = 255

        roi[roi < 0] = 0
        roi[(roi > 0) & (roi <= 127.5)] = 0
        roi[(roi > 127.5)] = 1
        seg[seg < 0] = 0
        seg[(seg > 0) & (seg <= 127.5)] = 0
        seg[(seg > 127.5)] = 1
        seg[roi == 0] = 0

        patch_size = 256
        x = np.arange(0, img.shape[0] - patch_size, 50)
        y = np.arange(0, img.shape[1] - patch_size, 50)
        xx, yy = np.meshgrid(x, y, sparse=True)
        pred_mask = np.zeros(img.shape)
        pred_mask[0:xx[0][-1], 0:yy[-1][0]] = 1

        patch_samples = np.ndarray((len(xx[0, ...]) * len(yy[..., 0]), patch_size, patch_size, 1), dtype='float32')
        k = 0
        for xi in xx[0, ...]:
            for yi in yy[..., 0]:
                x_min = int(xi)
                x_max = int(xi + patch_size)
                y_min = int(yi)
                y_max = int(yi + patch_size)
                patch_samples[k, ..., 0] = img[x_min:x_max, y_min:y_max]
                k += 1


        pred_sample = model.predict(patch_samples, batch_size=1)
        pred_seg = np.zeros(img.shape)
        weit_seg = np.zeros(img.shape)
        k = 0
        for xi in xx[0, ...]:
            for yi in yy[..., 0]:
                x_min = int(xi)
                x_max = int(xi + patch_size)
                y_min = int(yi)
                y_max = int(yi + patch_size)
                pred_seg[x_min:x_max, y_min:y_max] += pred_sample[k, ..., 0]
                weit_seg[x_min:x_max, y_min:y_max] += np.ones((patch_size, patch_size))
                k += 1
        weit_seg[weit_seg > 0] = 1 / weit_seg[weit_seg > 0]
        pred_res = np.multiply(pred_seg, weit_seg)
        #this is arbitrary, just so we can separate true positives and true negatives
        f2betascores = {}
        prscores = {}
        seg[seg == 0] = 10
        threshold = list(np.linspace(0.0, 1.0, 51))
        pscores = []
        rscores = []
        for i in range(len(threshold)):
            pred_res_copy = copy.deepcopy(pred_res)
            pred_res_copy[pred_res_copy >= threshold[i]] = 1
            pred_res_copy[pred_res_copy < threshold[i]] = 0
            pred_res_copy[roi == 0] = 0
            comparison = np.add(pred_res_copy, -1 * seg)

            true_positives = (comparison == 0).sum()
            false_positives = (comparison == -9).sum()
            false_negatives = (comparison == -1).sum()
            precision = true_positives / (false_positives + true_positives)
            recall = true_positives / (true_positives + false_negatives)
            f2beta = (1 + math.pow(2, 2)) * (precision * recall) / (math.pow(2, 2) * precision + recall)
            f2betascores[f2beta] = threshold[i]
            prscores[threshold[i]] = (recall, precision)
            pscores.append(precision)
            rscores.append(recall)


        maxthreshold = f2betascores[max(f2betascores.keys())]
        #saving f2betascores (f2beta -> threshold)
        np.save(os.path.join(result_save_folder, test_case, test_case + '-' + str(maxthreshold) + str(max(f2betascores.keys())) + 'f2betadict.npy'), np.array(f2betascores))
        print(str(maxthreshold))
        print(str(max(f2betascores.keys())))
        thresh_res = np.copy(pred_res)
        thresh_res[thresh_res >= maxthreshold] = 1
        thresh_res[thresh_res < maxthreshold] = 0

        #store previous value of precision and recall for trapezoid rule
        prevp = 0
        prevr = 0
        AUC = 0
        c = list(zip(rscores, pscores))
        #sort by recall
        sortedvals = sorted(c, key = lambda pair: pair[0])
        srecall, sprecision = zip(*sortedvals)

        for j in range(len(srecall)):
            deltax = srecall[j] - prevr
            yavg = (sprecision[j] + prevp) / 2
            AUC += deltax * yavg
            prevr = srecall[j]
            prevp = sprecision[j]

        #plotting and saving plot
        plt.scatter(rscores, pscores, label="AUC: " + str(AUC))
        plt.axis([0, 1, 0, 1])
        ax = plt.gca()
        ax.set_autoscale_on(False)
        plt.legend(loc='upper right')
        plt.xlabel('Recall')
        plt.ylabel('Precision')
        plt.title(test_case)
        plt.savefig(os.path.join(result_save_folder, test_case, test_case + 'prcurve.png'))
        plt.close()
        #saving images
        Image.fromarray(pred_res * 255).convert('RGB').save(os.path.join(result_save_folder, test_case, test_case + '_cnn.png'))
        Image.fromarray(seg * 255).convert('RGB').save(os.path.join(result_save_folder, test_case, test_case + '_seg.png'))
        Image.fromarray(img).convert('RGB').save(os.path.join(result_save_folder, test_case, test_case + '_img.png'))
        Image.fromarray(pred_mask * 255).convert('RGB').save(os.path.join(result_save_folder, test_case, test_case + '_cnn_mask.png'))
        Image.fromarray(thresh_res * 255).convert('RGB').save(os.path.join(result_save_folder, test_case, test_case + '_cnn_thresh.png'))
    return

#covert models to matlab
def convert_to_mat():
    save_folder = str(os.path.join('~', 'MATLAB Add-Ons', 'Apps', 'LeafVeinCNN', 'LeafVeinAnalysis', 'cnn_model', 'Matlab models'))
    print(save_folder)
    model_location = str(os.getcwd()) + '/' + str(os.path.join('models'))
    print(model_location)
    eng = matlab.engine.start_matlab()
    eng.convertCNN(nargout=0)

#calculate shannon entropy
def calculate_shannon_entropy():
    image_folder = os.path.join(os.getcwd(), 'images', 'all_images')
    image_list = [x for x in os.listdir(image_folder) if x[0] != '.']
    for img_name in image_list:
        img_file = img_name + '_img.png'
        #extracting green channel
        img = Image.open(os.path.join(image_folder, img_name, img_file)).convert("RGB")
        red, img, blue = img.split()
        img = np.array(img)
        entropy = skimage.measure.shannon_entropy(img)
        print(img_name + ": " + str(entropy))

#get best models after training
def getbestmodels():
    folds = os.listdir(os.path.join(os.getcwd(), 'folds'))
    if not os.path.exists('models'):
        os.mkdir('models')
    for fold in folds:
        exp_folder = os.listdir(os.path.join(os.getcwd(), 'folds', fold, 'results'))
        for exp in range(1, 1 + len(exp_folder)):
            exp = 'exp' + str(exp)
            cwd = os.getcwd()
            #grab model
            #divide test image into patches
            #predict on test image
            #show output
            if os.path.exists(os.path.join(cwd, 'folds', fold, 'results', exp, 'max_dice.npy')):
                file_arr = np.load(os.path.join(cwd, 'folds', fold, 'results', exp, 'max_dice.npy'))
                filename = file_arr[0]
                shutil.copy(os.path.join(cwd, 'folds', fold, 'model', exp, filename), os.path.join(cwd, 'models', filename))

#visualize data with T-SNE
def maketsne(pathes):
    numerator = 0
    patch_size = 256
    patches_per_image = 32
    trn_count = patches_per_image - numerator
    cwd = os.getcwd()

    trn_imgs = [os.listdir(x) for x in pathes]

    numdata = len(pathes)

    #call prep if data does not already exist
    #prep('f_tsne', 'exp1', trn_imgs[1], 0)
    #for i in range(numdata):
    #    print(len(trn_imgs[i]))
    #    prep('f_tsne', 'exp1', trn_imgs[i], 0)
    print("finished prep")
    data_folder = [os.path.join(x + "/../../..", 'folds', 'f_tsne', 'preprocessed_data', 'exp1', 'training_data') for x in pathes]
    datasize = [len(x) for x in trn_imgs]
    trn_img_data_stack = np.ndarray((sum(datasize) * trn_count, patch_size, patch_size), dtype='float32')
    patch_vectors = np.ndarray((sum(datasize) * trn_count, patch_size * patch_size), dtype='float32') #-480 because of memory issues


    trn_img_data = [np.load(os.path.join(x, 'img_trn_data.npy')) for x in data_folder]
    print("loaded datasets")

    count = 0
    for i in range(numdata):
        print(i)
        for j in range(len(trn_imgs[i]) * trn_count):
            trn_img_data_stack[count] = np.resize(trn_img_data[i][j], (patch_size, patch_size))
            count += 1

    print('processed patches')
    row = 0
    count = 0
    #removed_indices = random.sample(range(patch_vectors.shape[0]), 480)
    for patch in trn_img_data_stack:
        #if count not in removed_indices:
        patch_vectors[row, :] = np.resize(patch, (patch_size * patch_size))
        row += 1
        #count += 1

    print(patch_vectors.shape)


    print('reshaped input array')
    # Do PCA and keep 50 dimensions
    patch_vectors = patch_vectors - patch_vectors.mean(axis=0)
    print("demeaned")
    U, s, V = linalg.svd(patch_vectors, full_matrices=False, lapack_driver='gesvd')
    print("SVD completed")
    X50 = np.dot(U, np.diag(s))[:,:50]
    print('applied PCA')
    #some nice colors
    sett = np.array([[106,174,148],[109,73,204],[114,181,71],[201,73,190],[201,159,71],
                    [78,48,115],[210,78,50],[108,150,196],[203,73,114],[74,96,48],
                    [179,128,197],[152,98,57],[54,64,74],[191,145,146],[98,41,52]])
    col = sett[0:numdata]
    colors = np.ndarray((patch_vectors.shape[0], 3), dtype='float32')
    count = 0
    for k in range(len(col)):
        for i in range(datasize[k]):
            for j in range(patches_per_image):
                colors[count] = col[k]
                count += 1


    print('plotting')
    print(X50.shape)
    #tsne and plotting
    Z = fast_tsne(X50, late_exag_coeff=2, perplexity_list=[3,15,30,50])
    print(Z.shape)

    plt.figure(figsize=(4,4))
    plt.axis('equal')
    plt.scatter(Z[:,0], Z[:,1], c = colors/255.0, s=16, edgecolors='none')
    sns.despine()
    plt.tight_layout()
    plt.show()

#predict using models in a folder
def predict(folder, fold, exp, model_name, patch_size):
    cwd = os.getcwd()
    #grab model
    #print("hi")
    #divide test image into patches
    #predict on test image
    #show output
    test_folder = os.path.join(cwd, folder)
    test_sample = os.listdir(test_folder)
    result_save_folder = folder
    #we probably have to make an array of these

    #getting model location  from epoch
    model_location = os.path.join(os.getcwd(), 'folds', fold, 'model', exp, model_name)
    # list_of_files = glob.glob(model_folder + '/*') # * means all if need specific format then *.csv
    # if epoch < 10 and epoch != 0:
    #     name = '0' + str(epoch)
    # else:
    #     name = str(epoch)
    # model_location = [x for x in list_of_files if (name + '-') in x]

    #model = load_model(model_location[-1], compile=False)

    #model_location= os.path.join(cwd, 'f11-256-weights.24-0.383.h5') #new
    # if patch_size == 512:
    #     model_location = os.path.join(cwd, 'f14-512-weights.30-0.368.h5')
    # if patch_size == 256:
    #     model_location = os.path.join(cwd, 'f11-256-weights.24-0.383.h5')
    # elif patch_size == 1024:
    #     model_location = os.path.join(cwd, 'f15-1024-weights.31-0.316.h5')
    model = load_model(model_location, compile=False) #new

    count = 0
    for test_case in test_sample:
        if '_img.png' not in test_case:
            continue
        count += 1
        print('Test case: ' + test_case + 'Progress: ' + str(count) + ' / ' + str(len(test_sample)))

        # #get vein density stats
        # ps = openpyxl.load_workbook(os.path.join(test_folder, 'results_table_epoch_' + str(epoch) + '.xlsx'))
        # sheet = ps['Sheet1']
        # avg_vd = sheet['H2'].value
        # avg_loopiness = sheet['AH2'].value



        # original image
        img = Image.open(os.path.join(test_folder, test_case)).convert("L")
        img = ImageOps.autocontrast(img, cutoff = 0.01)
        img = np.array(img)

        roi_test_sample = test_case[:-8] + '_roi.png'
        # region of interest
        roi = Image.open(os.path.join(cwd, test_folder, roi_test_sample)).convert("L")
        roi = np.array(roi)
        roi[roi < 0] = 0 #2
        roi[(roi > 0) & (roi <= 127.5)] = 0 #2
        roi[(roi > 127.5)] = 2 #0


        x = np.arange(0, img.shape[0] - patch_size, patch_size // 4)
        y = np.arange(0, img.shape[1] - patch_size, patch_size // 4)
        xx, yy = np.meshgrid(x, y, sparse=True)
        pred_mask = np.zeros(img.shape)
        pred_mask[0:xx[0][-1], 0:yy[-1][0]] = 1

        patch_samples = np.ndarray((len(xx[0, ...]) * len(yy[..., 0]), patch_size, patch_size, 1), dtype='float32')
        pred_sample = np.ndarray((len(xx[0, ...]) * len(yy[..., 0]), patch_size, patch_size, 1), dtype='float32')

        k = 0
        pred_seg = np.zeros(img.shape)
        weit_seg = np.zeros(img.shape)
        for xi in xx[0, ...]:
            for yi in yy[..., 0]:
                x_min = int(xi)
                x_max = int(xi + patch_size)
                y_min = int(yi)
                y_max = int(yi + patch_size)
                patch_samples[k, ..., 0] = img[x_min:x_max, y_min:y_max]
                pred_sample[k, ...] = model.predict(patch_samples[np.newaxis, k, ...])
                pred_seg[x_min:x_max, y_min:y_max] += np.squeeze(pred_sample[k, ...])
                weit_seg[x_min:x_max, y_min:y_max] += np.ones((patch_size, patch_size))
                k += 1

        if voting:
            pred_res = np.zeros(img.shape)
            pred_res[pred_seg > (weit_seg / 2)] = 1
            pred_res[pred_seg < (weit_seg / 2)] = 0
        #print("averaging")
        # averaging
        else:
            weit_seg = 1 / weit_seg
            pred_res = np.multiply(pred_seg, weit_seg)



        #saving images
        Image.fromarray(pred_res * 255).convert('RGB').save(os.path.join(result_save_folder, test_case + '_cnn.png'))
    return

def convert_roi_to_grayscale(folder):
    roi_name = folder + '_roi.png'
    roi = Image.open(os.path.join(os.getcwd(), folder, roi_name)).convert("L")
    roi.save(os.path.join(folder, roi_name))
    return

def convert_img_to_grayscale(folder):
    img_name = folder + '_img.png'
    # get all folders inside the folder and get just the ones that contain the img_name in them
    # will be all the window size/threshold combos

    for leaf in os.listdir(os.path.join(os.getcwd(), folder, "den512")):
        # get image and grayscale it
        if 'png' == leaf[-3:]:
            location = os.path.join(os.getcwd(), folder, "den512", leaf)
            img = Image.open(location).convert("L")
            img.save(location)
    return


def predict_with_vd_thresholding(predict_folder, output_folder, test_case, patch_size, avg_vd, sliding_window_length, voting, model_location):
    cwd = os.getcwd()
    #grab model
    #divide test image into patches
    #predict on test image
    #show output

    #we probably have to make an array of these

    #getting model location  from epoch
    #model_folder = os.path.join(os.getcwd(), 'folds', fold, 'model', 'exp1')
    # list_of_files = glob.glob(model_folder + '/*') # * means all if need specific format then *.csv
    # if epoch < 10 and epoch != 0:
    #     name = '0' + str(epoch)
    # else:
    #     name = str(epoch)
    # model_location = [x for x in list_of_files if (name + '-') in x]

    #model = load_model(model_location[-1], compile=False)

    #model_location= os.path.join(cwd, 'f11-256-weights.24-0.383.h5') #new
    # if patch_size == 512:
    #     model_location = os.path.join(cwd, 'f14-512-weights.30-0.368.h5')
    # if patch_size == 256:
    #     model_location = os.path.join(cwd, 'f11-256-weights.24-0.383.h5')
    # elif patch_size == 1024:
    #     model_location = os.path.join(cwd, 'f15-1024-weights.31-0.316.h5')
    model = load_model(model_location, compile=False) #new


    # #get vein density stats
    # ps = openpyxl.load_workbook(os.path.join(test_folder, 'results_table_epoch_' + str(epoch) + '.xlsx'))
    # sheet = ps['Sheet1']
    # avg_vd = sheet['H2'].value
    # avg_loopiness = sheet['AH2'].value



    # original image
    img = Image.open(os.path.join(os.getcwd(), predict_folder, test_case)).convert("L")
    img = ImageOps.autocontrast(img, cutoff = 0.01)
    img = np.array(img)
    for elem in img.shape:
        if elem < 256:
            print("the dimensions of the image are too small in one or both dimensions - please check and resize image if necessary")
            return
    roi_test_sample = test_case[:-8] + '_roi.png'
    # region of interest
    roi = Image.open(os.path.join(os.getcwd(), predict_folder, roi_test_sample)).convert("L")
    roi = np.array(roi)
    roi[roi < 0] = 0 #2
    roi[(roi > 0) & (roi <= 127.5)] = 0 #2
    roi[(roi > 127.5)] = 2 #0


    x = np.arange(0, img.shape[0] - patch_size, patch_size // 4)
    y = np.arange(0, img.shape[1] - patch_size, patch_size // 4)
    xx, yy = np.meshgrid(x, y, sparse=True)

    #patch_samples = np.ndarray((len(xx[0, ...]) * len(yy[..., 0]), patch_size, patch_size, 1), dtype='float32')
    #pred_sample = np.ndarray((len(xx[0, ...]) * len(yy[..., 0]), patch_size, patch_size, 1), dtype='float32')

    k = 0
    pred_seg = np.zeros(img.shape)
    weit_seg = np.zeros(img.shape)
    for xi in xx[0, ...]:
        for yi in yy[..., 0]:
            x_min = int(xi)
            x_max = int(xi + patch_size)
            y_min = int(yi)
            y_max = int(yi + patch_size)
            patch_sample = img[x_min:x_max, y_min:y_max]
            # can potentially batch this
            pred_sample = model.predict(patch_sample[np.newaxis, ...])
            pred_seg[x_min:x_max, y_min:y_max] += np.squeeze(pred_sample)
            weit_seg[x_min:x_max, y_min:y_max] += np.ones((patch_size, patch_size))
            k += 1


    if voting:
        pred_res = np.zeros(img.shape)
        pred_res[pred_seg > (weit_seg / 2)] = 1
        pred_res[pred_seg < (weit_seg / 2)] = 0
    #print("averaging")
    # averaging
    else:
        weit_seg = 1 / weit_seg
        pred_res = np.multiply(pred_seg, weit_seg)
        del pred_seg

    x = np.arange(0, img.shape[0] - sliding_window_length, sliding_window_length // 4)
    y = np.arange(0, img.shape[1] - sliding_window_length, sliding_window_length // 4)
    xx, yy = np.meshgrid(x, y, sparse=True)

    thresh_seg = np.zeros(img.shape)
    weit_seg = np.zeros(img.shape)
    for xi in xx[0, ...]:
        #gc.collect()
        for yi in yy[..., 0]:
            x_min = int(xi)
            x_max = int(xi + sliding_window_length)
            y_min = int(yi)
            y_max = int(yi + sliding_window_length)
            #find best pred_sample using per patch thresholding off of seg vein density
            pred_window = copy.deepcopy(pred_res[x_min:x_max, y_min:y_max])
            threshold = list(np.linspace(0.0, 1.0, 100))
            vein_densities = []
            #threshold loop
            roi_window = roi[x_min:x_max, y_min:y_max]
            if np.count_nonzero(roi_window) > 0:
                for i in range(len(threshold)):
                    #calculate vd for patch and add to array
                    thresh_window = copy.deepcopy(pred_window)
                    thresh_window[thresh_window >= threshold[i]] = 1
                    thresh_window[thresh_window < threshold[i]] = 0

                    window_vd = get_vd_of_thresholded_patch(thresh_window, roi_window)

                    vein_densities.append(window_vd)
                #take patch with vd closest to the avg_vd
                #print(vein_densities)
                min_diff = 1e20
                best_threshold_i = 0
                for i in range(len(vein_densities)):
                    curr_vein_density = vein_densities[i]
                    vd_diff = np.abs(curr_vein_density - avg_vd)
                    if vd_diff < min_diff:
                        min_diff = vd_diff
                        best_threshold_i = i
                #print(f"threshold: {threshold[best_threshold_i]}")
                pred_window[pred_window >= threshold[best_threshold_i]] = 1
                pred_window[pred_window < threshold[best_threshold_i]] = 0
                pred_window[roi_window == 0] = 0

                thresh_seg[x_min:x_max, y_min:y_max] += pred_window
                weit_seg[x_min:x_max, y_min:y_max] += np.ones((sliding_window_length, sliding_window_length))
                #calculate vein density
                #print(f"vein density: {get_vd_of_thresholded_patch(pred_patch, roi_patch)}")
                #pred_patch[roi == 0] = 0
                # if np.count_nonzero(pred_patch) > 2000:
                #     roi_img = Image.fromarray(roi_patch * 255 // 2).convert('RGB')
                #     plt.imshow(roi_img)
                #     plt.show()
                #     pred_img = Image.fromarray(pred_patch * 255).convert('RGB')
                #     plt.imshow(pred_img)
                #     plt.show()

    if voting:
        pred_res = np.zeros(img.shape)
        pred_res[pred_seg > (weit_seg / 2)] = 1
        pred_res[pred_seg < (weit_seg / 2)] = 0
    else:
        # averaging
        weit_seg = 1 / weit_seg
        thresh_seg = np.multiply(thresh_seg, weit_seg)

    #saving images
    scaled_thresh_seg = thresh_seg * 255
    result_save_folder = output_folder
    Image.fromarray(scaled_thresh_seg).convert('L').save(os.path.join(result_save_folder, test_case + '_cnn' + f"_{sliding_window_length}_" + str(avg_vd) + '.png'))

    #save vein overlaid image
    C = np.dstack((scaled_thresh_seg, img, scaled_thresh_seg))
    Image.fromarray(C.astype(np.uint8)).convert('RGB').save(os.path.join(result_save_folder, test_case + '_cnn' + f"_{sliding_window_length}_" + str(avg_vd) + '_overlay' +  '.png'))
    return


def borneo_preprocess(location):
    #iterate through each image folder
    #get the image and segs and downsample by 23.8 then upsample
    #save in images folder of lowresleafcnn15
    cwd = os.getcwd()

    data_folder = os.path.join(os.getcwd(), location)
    image_folders = os.listdir(data_folder)
    result_save_folder = os.path.join(os.getcwd(), "images", "all_images")
    for image in image_folders:
        img = np.array(Image.open(os.path.join(data_folder, image, image + "_img.png")).convert("L"))
        seg = np.array(Image.open(os.path.join(data_folder, image, image + "_cnn_1.png")).convert("L"))
        roi = np.array(Image.open(os.path.join(data_folder, image, image + "_mask.png")).convert("L"))
        PIL_img = Image.fromarray(img).convert("L")
        img_width, img_height = PIL_img.size
        PIL_seg = Image.fromarray(seg).convert("L")
        seg_width, seg_height = PIL_seg.size
        PIL_roi = Image.fromarray(roi).convert("L")
        roi_width, roi_height = PIL_roi.size

        PIL_img = PIL_img.resize(((int) (img_width // 8), (int) (img_height // 8)))
        # PIL_seg = PIL_seg.resize(((int) (seg_width), (int) (seg_height)))
        #PIL_roi = PIL_roi.resize(((int) (roi_width), (int) (roi_height)))

        PIL_img = PIL_img.resize(((int) (img_width), (int) (img_height)))
       # PIL_seg = PIL_seg.resize(((int) (seg_width), (int) (seg_height)))
        #PIL_roi = PIL_roi.resize(((int) (roi_width), (int) (roi_height)))

        if not (os.path.isdir(os.path.join(result_save_folder, image))):
            os.mkdir(os.path.join(result_save_folder, image))

        PIL_img.convert('RGB').save(os.path.join(result_save_folder, image, image + '_img.png'))
        PIL_seg.convert('RGB').save(os.path.join(result_save_folder, image, image + '_seg.png'))
        PIL_roi.convert('RGB').save(os.path.join(result_save_folder, image, image + '_roi.png'))


def get_vd_of_thresholded_patch(thresh_patch, roi_patch):
    white, black = 0, 0
    diff = roi_patch - thresh_patch
    white = np.count_nonzero(diff == 1)
    black = np.count_nonzero(diff == 2)

    if black == 0 and white == 0:
        black = 1
    patch_vd = white / (black + white)

    return patch_vd

def display_vd_thresholds(folder, fold, patch_size, epoch, avg_vd, sliding_window_length):
    cwd = os.getcwd()
    #grab model
    #divide test image into patches
    #predict on test image
    #show output
    test_folder = os.path.join(cwd, folder)
    test_sample = os.listdir(test_folder)
    result_save_folder = folder
    #we probably have to make an array of these

    #getting model location  from epoch
    #model_folder = os.path.join(os.getcwd(), 'folds', fold, 'model', 'exp1')
    # list_of_files = glob.glob(model_folder + '/*') # * means all if need specific format then *.csv
    # if epoch < 10 and epoch != 0:
    #     name = '0' + str(epoch)
    # else:
    #     name = str(epoch)
    # model_location = [x for x in list_of_files if (name + '-') in x]

    #model = load_model(model_location[-1], compile=False)

    #model_location= os.path.join(cwd, 'f11-256-weights.24-0.383.h5') #new
    if patch_size == 512:
        model_location = os.path.join(cwd, 'f14-512-weights.30-0.368.h5')
    if patch_size == 256:
        model_location = os.path.join(cwd, 'f11-256-weights.24-0.383.h5')
    elif patch_size == 1024:
        model_location = os.path.join(cwd, 'f15-1024-weights.31-0.316.h5')
    model = load_model(model_location, compile=False) #new

    count = 0
    for test_case in test_sample:
        if '_img.png' not in test_case:
            continue
        count += 1
        print('Test case: ' + test_case + 'Progress: ' + str(count) + ' / ' + str(len(test_sample)))

        # #get vein density stats
        # ps = openpyxl.load_workbook(os.path.join(test_folder, 'results_table_epoch_' + str(epoch) + '.xlsx'))
        # sheet = ps['Sheet1']
        # avg_vd = sheet['H2'].value
        # avg_loopiness = sheet['AH2'].value



        # original image
        img = Image.open(os.path.join(test_folder, test_case)).convert("L")
        img = ImageOps.autocontrast(img, cutoff = 0.01)
        img = np.array(img)

        roi_test_sample = test_case[:-8] + '_roi.png'
        # region of interest
        roi = Image.open(os.path.join(cwd, test_folder, roi_test_sample)).convert("L")
        roi = np.array(roi)
        # If Satvik makes the roi, flip
        roi[roi < 0] = 0 #2
        roi[(roi > 0) & (roi <= 127.5)] = 0 #2
        roi[(roi > 127.5)] = 2 #0


        x = np.arange(0, img.shape[0] - patch_size, patch_size // 4)
        y = np.arange(0, img.shape[1] - patch_size, patch_size // 4)
        xx, yy = np.meshgrid(x, y, sparse=True)
        pred_mask = np.zeros(img.shape)
        pred_mask[0:xx[0][-1], 0:yy[-1][0]] = 1

        patch_samples = np.ndarray((len(xx[0, ...]) * len(yy[..., 0]), patch_size, patch_size, 1), dtype='float32')
        pred_sample = np.ndarray((len(xx[0, ...]) * len(yy[..., 0]), patch_size, patch_size, 1), dtype='float32')

        k = 0
        pred_seg = np.zeros(img.shape)
        weit_seg = np.zeros(img.shape)
        for xi in xx[0, ...]:
            for yi in yy[..., 0]:
                x_min = int(xi)
                x_max = int(xi + patch_size)
                y_min = int(yi)
                y_max = int(yi + patch_size)
                patch_samples[k, ..., 0] = img[x_min:x_max, y_min:y_max]
                pred_sample[k, ...] = model.predict(patch_samples[np.newaxis, k, ...])
                pred_seg[x_min:x_max, y_min:y_max] += np.squeeze(pred_sample[k, ...])
                weit_seg[x_min:x_max, y_min:y_max] += np.ones((patch_size, patch_size))
                k += 1

        # voting
        #pred_res = np.zeros(img.shape)
        #pred_res[pred_seg > (weit_seg / 2)] = 1
        #pred_res[pred_seg < (weit_seg / 2)] = 0
        print("averaging")
        # averaging
        weit_seg = 1 / weit_seg
        pred_res = np.multiply(pred_seg, weit_seg)

        x = np.arange(0, img.shape[0] - sliding_window_length, sliding_window_length // 4)
        y = np.arange(0, img.shape[1] - sliding_window_length, sliding_window_length // 4)
        xx, yy = np.meshgrid(x, y, sparse=True)

        thresh_seg = np.zeros(img.shape)
        weit_seg = np.zeros(img.shape)
        for xi in xx[0, ...]:
            for yi in yy[..., 0]:
                x_min = int(xi)
                x_max = int(xi + sliding_window_length)
                y_min = int(yi)
                y_max = int(yi + sliding_window_length)
                #find best pred_sample using per patch thresholding off of seg vein density
                pred_window = copy.deepcopy(pred_res[x_min:x_max, y_min:y_max])
                threshold = list(np.linspace(0.0, 1.0, 100))
                vein_densities = []
                #threshold loop
                roi_window = roi[x_min:x_max, y_min:y_max]

                if np.count_nonzero(roi_window) > 0:
                    for i in range(len(threshold)):
                        #calculate vd for patch and add to array
                        thresh_window = copy.deepcopy(pred_window)
                        thresh_window[thresh_window >= threshold[i]] = 1
                        thresh_window[thresh_window < threshold[i]] = 0

                        window_vd = get_vd_of_thresholded_patch(thresh_window, roi_window)

                        vein_densities.append(window_vd)

                    #take patch with vd closest t   o the avg_vd
                    #print(vein_densities)
                    min_diff = 1e20
                    best_threshold_i = 0
                    for i in range(len(vein_densities)):
                        curr_vein_density = vein_densities[i]
                        vd_diff = np.abs(curr_vein_density - avg_vd)
                        if vd_diff < min_diff:
                            min_diff = vd_diff
                            best_threshold_i = i
                    #print(f"threshold: {threshold[best_threshold_i]}")
                    pred_window[:] = threshold[best_threshold_i]
                    #pred_window[pred_window < threshold[best_threshold_i]] = 0
                    pred_window[roi_window == 0] = 0

                    thresh_seg[x_min:x_max, y_min:y_max] += pred_window
                    weit_seg[x_min:x_max, y_min:y_max] += np.ones((sliding_window_length, sliding_window_length))

                    #calculate vein density
                    #print(f"vein density: {get_vd_of_thresholded_patch(pred_patch, roi_patch)}")
                    #pred_patch[roi == 0] = 0
                    # if np.count_nonzero(pred_patch) > 2000:
                    #     roi_img = Image.fromarray(roi_patch * 255 // 2).convert('RGB')
                    #     plt.imshow(roi_img)
                    #     plt.show()
                    #     pred_img = Image.fromarray(pred_patch * 255).convert('RGB')
                    #     plt.imshow(pred_img)
                    #     plt.show()

        # voting
        #pred_res = np.zeros(img.shape)
        #pred_res[pred_seg > (weit_seg / 2)] = 1
        #pred_res[pred_seg < (weit_seg / 2)] = 0
        print("averaging")
        # averaging
        weit_seg = 1 / weit_seg
        thresh_seg = np.multiply(thresh_seg, weit_seg)

        #saving images
        Image.fromarray(thresh_seg * 255).convert('RGB').save(os.path.join(result_save_folder, f"den{patch_size}", test_case + 'thresh_viz' + f"_{sliding_window_length}_" + str(avg_vd) + '.png'))
    return

def display_quality_regions(folder, fold, patch_size, epoch, list_of_vds, sliding_window_length):
    cwd = os.getcwd()
    #grab model
    print("hi")
    #divide test image into patches
    #predict on test image
    #show output
    test_folder = os.path.join(cwd, folder)
    test_sample = os.listdir(test_folder)
    result_save_folder = folder
    #we probably have to make an array of these

    #getting model location  from epoch
    #model_folder = os.path.join(os.getcwd(), 'folds', fold, 'model', 'exp1')
    # list_of_files = glob.glob(model_folder + '/*') # * means all if need specific format then *.csv
    # if epoch < 10 and epoch != 0:
    #     name = '0' + str(epoch)
    # else:
    #     name = str(epoch)
    # model_location = [x for x in list_of_files if (name + '-') in x]

    #model = load_model(model_location[-1], compile=False)

    #model_location= os.path.join(cwd, 'f11-256-weights.24-0.383.h5') #new
    if patch_size == 512:
        model_location = os.path.join(cwd, 'f14-512-weights.30-0.368.h5')
    if patch_size == 256:
        model_location = os.path.join(cwd, 'f11-256-weights.24-0.383.h5')
    elif patch_size == 1024:
        model_location = os.path.join(cwd, 'f15-1024-weights.31-0.316.h5')
    model = load_model(model_location, compile=False) #new

    count = 0
    for test_case in test_sample:
        if '_img.png' not in test_case:
            continue
        count += 1
        print('Test case: ' + test_case + 'Progress: ' + str(count) + ' / ' + str(len(test_sample)))

        # #get vein density stats
        # ps = openpyxl.load_workbook(os.path.join(test_folder, 'results_table_epoch_' + str(epoch) + '.xlsx'))
        # sheet = ps['Sheet1']
        # avg_vd = sheet['H2'].value
        # avg_loopiness = sheet['AH2'].value



        # original image
        img = Image.open(os.path.join(test_folder, test_case)).convert("L")
        img = ImageOps.autocontrast(img, cutoff = 0.01)
        img = np.array(img)

        roi_test_sample = test_case[:-8] + '_roi.png'
        # region of interest
        roi = Image.open(os.path.join(cwd, test_folder, roi_test_sample)).convert("L")
        roi = np.array(roi)
        roi[roi < 0] = 2
        roi[(roi > 0) & (roi <= 127.5)] = 2
        roi[(roi > 127.5)] = 0


        x = np.arange(0, img.shape[0] - patch_size, patch_size // 4)
        y = np.arange(0, img.shape[1] - patch_size, patch_size // 4)
        xx, yy = np.meshgrid(x, y, sparse=True)
        pred_mask = np.zeros(img.shape)
        pred_mask[0:xx[0][-1], 0:yy[-1][0]] = 1

        patch_samples = np.ndarray((len(xx[0, ...]) * len(yy[..., 0]), patch_size, patch_size, 1), dtype='float32')
        pred_sample = np.ndarray((len(xx[0, ...]) * len(yy[..., 0]), patch_size, patch_size, 1), dtype='float32')

        k = 0
        pred_seg = np.zeros(img.shape)
        weit_seg = np.zeros(img.shape)
        for xi in xx[0, ...]:
            for yi in yy[..., 0]:
                x_min = int(xi)
                x_max = int(xi + patch_size)
                y_min = int(yi)
                y_max = int(yi + patch_size)
                patch_samples[k, ..., 0] = img[x_min:x_max, y_min:y_max]
                pred_sample[k, ...] = model.predict(patch_samples[np.newaxis, k, ...])
                pred_seg[x_min:x_max, y_min:y_max] += np.squeeze(pred_sample[k, ...])
                weit_seg[x_min:x_max, y_min:y_max] += np.ones((patch_size, patch_size))
                k += 1

        # voting
        #pred_res = np.zeros(img.shape)
        #pred_res[pred_seg > (weit_seg / 2)] = 1
        #pred_res[pred_seg < (weit_seg / 2)] = 0
        print("averaging")
        # averaging
        weit_seg = 1 / weit_seg
        pred_res = np.multiply(pred_seg, weit_seg)

        x = np.arange(0, img.shape[0] - sliding_window_length, sliding_window_length // 4)
        y = np.arange(0, img.shape[1] - sliding_window_length, sliding_window_length // 4)
        xx, yy = np.meshgrid(x, y, sparse=True)

        thresh_seg = np.zeros(img.shape)
        weit_seg = np.zeros(img.shape)
        for xi in xx[0, ...]:
            for yi in yy[..., 0]:
                x_min = int(xi)
                x_max = int(xi + sliding_window_length)
                y_min = int(yi)
                y_max = int(yi + sliding_window_length)
                #find best pred_sample using per patch thresholding off of seg vein density
                pred_window = copy.deepcopy(pred_res[x_min:x_max, y_min:y_max])
                threshold = list(np.linspace(0.0, 1.0, 100))
                vein_densities = []
                #threshold loop
                roi_window = roi[x_min:x_max, y_min:y_max]

                if np.count_nonzero(roi_window) > 0:
                    for i in range(len(threshold)):
                        #calculate vd for patch and add to array
                        thresh_window = copy.deepcopy(pred_window)
                        thresh_window[thresh_window >= threshold[i]] = 1
                        thresh_window[thresh_window < threshold[i]] = 0

                        window_vd = get_vd_of_thresholded_patch(thresh_window, roi_window)

                        vein_densities.append(window_vd)

                    #take patch with vd closest t   o the avg_vd
                    #print(vein_densities)
                    best_thresholds = []
                    for avg_vd in list_of_vds:
                        min_diff = 1e20
                        best_threshold_i = 0
                        for i in range(len(vein_densities)):
                            curr_vein_density = vein_densities[i]
                            vd_diff = np.abs(curr_vein_density - avg_vd)
                            if vd_diff < min_diff:
                                min_diff = vd_diff
                                best_threshold_i = i
                        best_thresholds.append(threshold[best_threshold_i])

                    std_dev = np.std(best_thresholds)
                    #print(f"threshold: {threshold[best_threshold_i]}")
                    pred_window[:] = std_dev
                    #pred_window[pred_window < threshold[best_threshold_i]] = 0
                    pred_window[roi_window == 0] = 0

                    thresh_seg[x_min:x_max, y_min:y_max] += pred_window
                    weit_seg[x_min:x_max, y_min:y_max] += np.ones((sliding_window_length, sliding_window_length))

                    #calculate vein density
                    #print(f"vein density: {get_vd_of_thresholded_patch(pred_patch, roi_patch)}")
                    #pred_patch[roi == 0] = 0
                    # if np.count_nonzero(pred_patch) > 2000:
                    #     roi_img = Image.fromarray(roi_patch * 255 // 2).convert('RGB')
                    #     plt.imshow(roi_img)
                    #     plt.show()
                    #     pred_img = Image.fromarray(pred_patch * 255).convert('RGB')
                    #     plt.imshow(pred_img)
                    #     plt.show()

        # voting
        #pred_res = np.zeros(img.shape)
        #pred_res[pred_seg > (weit_seg / 2)] = 1
        #pred_res[pred_seg < (weit_seg / 2)] = 0

        print("averaging")
        # averaging
        weit_seg = 1 / weit_seg
        thresh_seg = np.multiply(thresh_seg, weit_seg)

        maximum = np.amax(thresh_seg)
        thresh_seg = thresh_seg / maximum
        thresh_seg = 1 - thresh_seg

        #saving images
        Image.fromarray(thresh_seg * 255).convert('RGB').save(os.path.join(result_save_folder, f"den{patch_size}", test_case + '_std' + f"_{sliding_window_length}_" + str(avg_vd) + '.png'))
    return
